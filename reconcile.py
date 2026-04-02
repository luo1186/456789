"""
核心对账引擎 v10
关联键：采购单号 + SKU
对账逻辑：
  1. 行数校验：同一PO+SKU，收货单行数 vs 对账单行数
  2. 按行比对：对账单顺序不动，收货单按对账单顺序依次消耗配对
  3. 数量校验：收货单数量为正确值
  4. 单价校验：采购单单价为正确值（含税单价优先）
  5. 行金额校验：系统计算 收货单数量×采购单单价 vs 对账单行金额
  6. 总额校验：所有行金额加总 vs 对账单单据总额
"""
import os, json, datetime, traceback, re
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from sqlalchemy.orm import Session

from database import SessionLocal
import models

RED_FILL    = PatternFill("solid", fgColor="FFE0E0")
RED_FONT    = Font(color="CC0000", bold=True)
GREEN_FILL  = PatternFill("solid", fgColor="E0FFE8")
GREEN_FONT  = Font(color="006400")
HEAD_FILL   = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT   = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COL_KEYWORDS = {
    "采购单号": ["采购单号","关联单据号","订单号","采购编号","PO号","客户订单号码","客户订单号"],
    "品名":     ["品名","品名规格","品名/规格","物料名称","商品名称","货品名","名称","产品名"],
    "SKU":      ["SKU","sku","料号","货号","物料编码","商品编码"],
    "含税单价": ["含税单价","含税价","税含单价"],
    "单价":     ["单价","单价/桶","采购单价","结算单价","price","Price","unit_price","不含税单价"],
    "数量":     ["到货量","数量","收货数量","实收数量","结算数量","qty","Qty","quantity","通知收货量"],
    "行金额":   ["行金额","金额","小计","价税合计","含税金额","amount","总价","合价"],
    "单据总额": ["单据总额","总额","合计","发票金额","total","Total","合计金额"],
}
SUMMARY_KW = ["合计","汇总","小计","总计","合计（大写）"]


# ══════════════════════════════════════════════════════
#  任务入口
# ══════════════════════════════════════════════════════
def run(task_id: str, file_paths: dict, result_dir: str):
    db: Session = SessionLocal()
    try:
        task = db.query(models.Task).filter(models.Task.id == task_id).first()
        if not task:
            return
        task.status = "processing"
        db.commit()
        summary, result_path = _reconcile(task_id, file_paths, result_dir, task.name)
        task.status = "done"
        task.result_path = result_path
        task.result_summary = json.dumps(summary, ensure_ascii=False)
        task.finished_at = datetime.datetime.utcnow()
        db.commit()
    except Exception as e:
        tb = traceback.format_exc()
        task = db.query(models.Task).filter(models.Task.id == task_id).first()
        if task:
            task.status = "failed"
            task.error_msg = f"{str(e)}\n{tb}"
            task.finished_at = datetime.datetime.utcnow()
            db.commit()
    finally:
        db.close()


# ══════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════
def _to_str(v) -> str:
    if v is None: return ""
    try:
        if isinstance(v, float) and np.isnan(v): return ""
    except Exception: pass
    return str(v).strip()

def _to_float(v):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)): return None
        s = str(v).replace(",","").replace("¥","").replace("￥","").strip()
        if s in ("","nan","None","-"): return None
        return float(s)
    except Exception:
        return None

def _to_float_d(v, default=0.0):
    r = _to_float(v)
    return r if r is not None else default

def _clean_po_no(v) -> str:
    s = _to_str(v)
    m = re.match(r'([A-Za-z]+[0-9]+[A-Za-z0-9-]*)', s)
    return m.group(1) if m else s

def _clean_sku(v) -> str:
    if v is None: return ""
    try:
        if isinstance(v, float) and np.isnan(v): return ""
    except Exception: pass
    s = str(v).strip()
    if s in ("","nan","None"): return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ══════════════════════════════════════════════════════
#  智能读取
# ══════════════════════════════════════════════════════
def _find_header_row(all_rows: list, needed_keys: list) -> int:
    best_idx, best_score = 0, 0
    for ri, row in enumerate(all_rows[:40]):
        row_text = " ".join([str(v) for v in row])
        score = sum(
            1 for t in needed_keys
            if any(kw in row_text for kw in COL_KEYWORDS.get(t, [t]))
        )
        if score > best_score:
            best_score, best_idx = score, ri
    if best_idx + 1 < len(all_rows):
        if all_rows[best_idx][:5] == all_rows[best_idx + 1][:5]:
            best_idx += 1
    return best_idx

def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")
    def is_noise(row):
        vals = [str(v).strip() for v in row if str(v).strip() not in ("","nan","None")]
        if not vals: return True
        if any(any(kw in v for kw in SUMMARY_KW) for v in vals): return True
        return False
    df = df[~df.apply(is_noise, axis=1)].reset_index(drop=True)
    unnamed = [c for c in df.columns if str(c).startswith("Unnamed") and df[c].isna().all()]
    return df.drop(columns=unnamed)

def _smart_read(path: str, needed_keys: list) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        for enc in ("utf-8-sig","gbk","utf-8"):
            try:
                raw = pd.read_csv(path, encoding=enc, header=None, dtype=str)
                hi = _find_header_row(raw.fillna("").values.tolist(), needed_keys)
                df = pd.read_csv(path, encoding=enc, header=hi, dtype=str)
                return _clean_df(df)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"CSV编码识别失败：{os.path.basename(path)}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = [[str(v).strip() if v is not None else "" for v in row]
                for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not all_rows:
        raise ValueError(f"文件为空：{os.path.basename(path)}")
    hi = _find_header_row(all_rows, needed_keys)
    engine = "openpyxl" if ext in (".xlsx",".xlsm") else "xlrd"
    df = pd.read_excel(path, engine=engine, header=hi, dtype=str)
    return _clean_df(df)


# ══════════════════════════════════════════════════════
#  列名归一化
# ══════════════════════════════════════════════════════
def _normalize_cols(df: pd.DataFrame, targets: list) -> pd.DataFrame:
    rename, used = {}, set()
    for col in df.columns:
        col_s = str(col).strip()
        for target in targets:
            if target in used: continue
            if any(kw in col_s for kw in COL_KEYWORDS.get(target, [target])):
                rename[col] = target
                used.add(target)
                break
    return df.rename(columns=rename)

def _normalize_po_price_cols(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    tax_col = plain_col = None
    for col in df.columns:
        col_s = str(col).strip()
        if tax_col is None and any(kw in col_s for kw in COL_KEYWORDS["含税单价"]):
            tax_col = col; rename[col] = "含税单价"
        elif plain_col is None and col != tax_col and any(kw in col_s for kw in COL_KEYWORDS["单价"]):
            plain_col = col; rename[col] = "单价"
    return df.rename(columns=rename)

def _check_cols(df, required, name):
    missing = [c for c in required if c not in df.columns]
    if missing:
        avail = [c for c in df.columns if not str(c).startswith("Unnamed")]
        raise ValueError(f"【{name}】未能识别字段：{missing}\n文件实际列名：{avail}")

def _get_po_price(prow) -> tuple:
    tax   = _to_float(prow.get("含税单价")) if "含税单价" in prow.index else None
    plain = _to_float(prow.get("单价"))     if "单价"    in prow.index else None
    if tax   is not None and tax   > 0: return tax,   "含税单价"
    if plain is not None and plain > 0: return plain, "单价"
    return None, "未找到"


# ══════════════════════════════════════════════════════
#  核心对账逻辑
# ══════════════════════════════════════════════════════
def _reconcile(task_id, file_paths, result_dir, task_name):

    po_df   = _smart_read(file_paths["po"],   ["采购单号","SKU","含税单价","单价"])
    recv_df = _smart_read(file_paths["recv"], ["采购单号","SKU","数量"])
    stmt_df = _smart_read(file_paths["stmt"], ["采购单号","SKU","数量","单价","行金额"])

    po_df   = _normalize_po_price_cols(po_df)
    po_df   = _normalize_cols(po_df,   ["采购单号","品名","SKU"])
    recv_df = _normalize_cols(recv_df, ["采购单号","品名","SKU","数量"])
    stmt_df = _normalize_cols(stmt_df, ["采购单号","品名","SKU","数量","单价","行金额","单据总额"])

    if "含税单价" not in po_df.columns and "单价" not in po_df.columns:
        raise ValueError("【采购订单表】未找到单价，需包含「含税单价」或「单价」列")
    _check_cols(po_df,   ["采购单号","SKU"],                            "采购订单表")
    _check_cols(recv_df, ["采购单号","SKU","数量"],                     "收货单表")
    _check_cols(stmt_df, ["采购单号","SKU","数量","单价","行金额"],      "电子对账单")

    po_df   = po_df.copy()
    recv_df = recv_df.copy()
    stmt_df = stmt_df.copy()

    po_df["采购单号"]   = po_df["采购单号"].apply(_to_str).replace("",np.nan).ffill().apply(_clean_po_no)
    po_df["SKU"]        = po_df["SKU"].apply(_clean_sku)
    recv_df["采购单号"] = recv_df["采购单号"].apply(_clean_po_no)
    recv_df["SKU"]      = recv_df["SKU"].apply(_clean_sku)
    stmt_df["采购单号"] = stmt_df["采购单号"].apply(_clean_po_no)
    stmt_df["SKU"]      = stmt_df["SKU"].apply(_clean_sku)

    po_df   = po_df[(po_df["采购单号"]!="")   & (po_df["SKU"]!="")].reset_index(drop=True)
    recv_df = recv_df[(recv_df["采购单号"]!="") & (recv_df["SKU"]!="")].reset_index(drop=True)
    stmt_df = stmt_df[(stmt_df["采购单号"]!="") & (stmt_df["SKU"]!="")].reset_index(drop=True)

    has_total   = "单据总额" in stmt_df.columns
    details     = []
    anomalies   = 0
    anomaly_amt = 0.0
    total_rows  = 0

    # ── 按对账单原始行顺序逐行处理（不用groupby，保留原始顺序）─
    # recv_cursor：记录每个 采购单号+SKU 在收货单里已消耗到第几行
    recv_cursor = {}  # key: (po_no, sku) → 下一个可用的收货单行索引

    # 先预建收货单索引，按 采购单号+SKU 分组，保持收货单原始顺序
    recv_index_raw = {}  # key: (po_no, sku) → list of row Series（原始顺序）
    for _, rrow in recv_df.iterrows():
        key = (_to_str(rrow.get("采购单号","")), _to_str(rrow.get("SKU","")))
        if key not in recv_index_raw:
            recv_index_raw[key] = []
        recv_index_raw[key].append(rrow)

    # 按对账单里该组的数量顺序，重新排列收货单行
    # 对账单各组的数量顺序
    stmt_qty_order = {}  # key: (po_no, sku) → list of 数量（对账单顺序）
    for _, srow in stmt_df.iterrows():
        key = (_to_str(srow.get("采购单号","")), _to_str(srow.get("SKU","")))
        if key not in stmt_qty_order:
            stmt_qty_order[key] = []
        stmt_qty_order[key].append(_to_float_d(srow.get("数量", 0)))

    recv_index = {}
    for key, recv_rows in recv_index_raw.items():
        stmt_qtys = stmt_qty_order.get(key, [])
        if len(recv_rows) != len(stmt_qtys):
            # 行数不一致时保持原始顺序（后续行数校验会报错）
            recv_index[key] = recv_rows
            continue
        # 按对账单数量顺序重新排列收货单行
        recv_remaining = list(recv_rows)
        reordered = []
        for sq in stmt_qtys:
            # 找收货单中数量最接近 sq 的未使用行
            best_i = None
            best_diff = float("inf")
            for i, rrow in enumerate(recv_remaining):
                rq = _to_float_d(rrow.get("数量", 0))
                diff = abs(rq - sq)
                if diff < best_diff:
                    best_diff = diff
                    best_i = i
            if best_i is not None:
                reordered.append(recv_remaining.pop(best_i))
        # 如果有剩余（理论上不会，因为行数相等）
        reordered.extend(recv_remaining)
        recv_index[key] = reordered

    # 先做行数预校验：每个 采购单号+SKU 的收货单行数 vs 对账单行数
    stmt_count = {}
    for _, srow in stmt_df.iterrows():
        key = (_to_str(srow.get("采购单号","")), _to_str(srow.get("SKU","")))
        stmt_count[key] = stmt_count.get(key, 0) + 1

    row_count_anomaly = set()  # 记录行数异常的key，跳过后续比对
    for key, s_cnt in stmt_count.items():
        r_cnt = len(recv_index.get(key, []))
        if r_cnt != s_cnt:
            row_count_anomaly.add(key)

    # 按对账单原始行顺序逐行处理
    for _, srow in stmt_df.iterrows():
        po_no  = _to_str(srow.get("采购单号",""))
        sku    = _to_str(srow.get("SKU",""))
        item   = _to_str(srow.get("品名",""))
        s_qty  = _to_float_d(srow.get("数量",0))
        s_price= _to_float_d(srow.get("单价",0))
        s_amt  = _to_float_d(srow.get("行金额",0))
        key    = (po_no, sku)

        if not po_no or not sku:
            continue
        total_rows += 1

        # 获取该组采购单价
        prow = po_df[(po_df["采购单号"]==po_no) & (po_df["SKU"]==sku)]
        p_price, price_src = (_get_po_price(prow.iloc[0]) if len(prow)>0 else (None,"未找到"))

        # 行数异常：整组报异常，不做明细比对
        if key in row_count_anomaly:
            r_cnt = len(recv_index.get(key, []))
            s_cnt = stmt_count[key]
            anomalies  += 1
            anomaly_amt += s_amt
            details.append({
                "po_no": po_no, "sku": sku, "item": item,
                "check_type": "行数异常",
                "recv_qty": None, "stmt_qty": s_qty,
                "po_price": p_price, "po_price_src": price_src,
                "stmt_price": s_price, "calc_amt": None, "stmt_amt": s_amt,
                "anomaly": True,
                "note": f"收货单{r_cnt}行 vs 对账单{s_cnt}行，行数不一致"
            })
            continue

        # 按顺序取下一条收货单记录（依次消耗）
        cursor = recv_cursor.get(key, 0)
        recv_list = recv_index.get(key, [])
        if cursor < len(recv_list):
            recv_cursor[key] = cursor + 1
            r_qty = _to_float_d(recv_list[cursor].get("数量", 0))
        else:
            r_qty = None

        errors = []; is_anomaly = False

        # 数量校验（收货单为准）
        if r_qty is None:
            errors.append("收货单未找到对应行"); is_anomaly = True
        elif abs(r_qty - s_qty) > 0.001:
            errors.append(f"数量差异：收货单{r_qty} vs 对账单{s_qty}"); is_anomaly = True

        # 单价校验（采购单为准）
        if p_price is None:
            errors.append(f"采购单未找到 采购单号[{po_no}]+SKU[{sku}]"); is_anomaly = True
        elif abs(p_price - s_price) > 0.001:
            errors.append(f"单价差异（{price_src}）：采购¥{p_price} vs 对账¥{s_price}"); is_anomaly = True

        # 行金额校验：系统计算 = 收货单数量 × 采购单单价
        if r_qty is not None and p_price is not None:
            calc_amt = round(r_qty * p_price, 2)
            if abs(calc_amt - s_amt) > 0.01:
                errors.append(f"行金额差异：系统计算¥{calc_amt} vs 对账单¥{s_amt}")
                is_anomaly = True
        else:
            calc_amt = None

        if is_anomaly:
            anomalies += 1
            anomaly_amt += s_amt

        details.append({
            "po_no": po_no, "sku": sku, "item": item,
            "check_type": "、".join(errors) if errors else "通过",
            "recv_qty": r_qty, "stmt_qty": s_qty,
            "po_price": p_price, "po_price_src": price_src,
            "stmt_price": s_price,
            "calc_amt": calc_amt, "stmt_amt": s_amt,
            "anomaly": is_anomaly,
            "note": "；".join(errors) if errors else "—"
        })

    # ── 总额校验 ─────────────────────────────────────────
    if has_total:
        for po_no in stmt_df["采购单号"].unique():
            po_rows    = stmt_df[stmt_df["采购单号"]==po_no]
            calc_total = sum(_to_float_d(r.get("行金额",0)) for _,r in po_rows.iterrows())
            stmt_total = _to_float(_to_str(po_rows["单据总额"].values[0])) if len(po_rows) else None
            if stmt_total and abs(calc_total - stmt_total) > 0.01:
                anomalies += 1
                anomaly_amt += abs(calc_total - stmt_total)
                details.append({
                    "po_no": po_no, "sku": "—", "item": "【总额校验】",
                    "check_type": "总额异常",
                    "recv_qty": None, "stmt_qty": None,
                    "po_price": None, "po_price_src": "—",
                    "stmt_price": None, "calc_amt": calc_total, "stmt_amt": stmt_total,
                    "anomaly": True,
                    "note": f"明细合计¥{calc_total:.2f} vs 单据总额¥{stmt_total:.2f}，差额¥{calc_total-stmt_total:.2f}"
                })

    os.makedirs(result_dir, exist_ok=True)
    result_path = os.path.join(result_dir, f"{task_id}_result.xlsx")
    _write_excel(result_path, task_name, details, po_df, recv_df, stmt_df)

    return {
        "total": total_rows, "passed": total_rows - anomalies,
        "anomalies": anomalies, "anomaly_amt": round(anomaly_amt, 2),
        "details": details
    }, result_path


# ══════════════════════════════════════════════════════
#  生成结果 Excel
# ══════════════════════════════════════════════════════
def _write_excel(path, task_name, details, po_df, recv_df, stmt_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "对账明细（含批注）"

    headers    = ["采购单号","SKU","品名","校验结果",
                  "收货单数量","对账单数量",
                  "采购单单价","单价来源","对账单单价",
                  "系统计算行金额","对账单行金额",
                  "异常说明","状态"]
    col_widths = [16,10,22,16,12,12,12,10,12,14,14,42,10]

    for ci,(h,w) in enumerate(zip(headers,col_widths),1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri,d in enumerate(details,2):
        row_data = [
            d["po_no"], d["sku"], d["item"], d["check_type"],
            d["recv_qty"], d["stmt_qty"],
            d["po_price"], d.get("po_price_src","—"), d["stmt_price"],
            d["calc_amt"], d["stmt_amt"],
            d["note"], "⚠ 异常" if d["anomaly"] else "✅ 通过"
        ]
        for ci,val in enumerate(row_data,1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
            c.fill = RED_FILL if d["anomaly"] else GREEN_FILL
            c.font = RED_FONT if d["anomaly"] else GREEN_FONT
        if d["anomaly"] and d["note"] and d["note"] != "—":
            nc = ws.cell(row=ri, column=12)
            cmt = Comment(d["note"],"ReconCore"); cmt.width=320; cmt.height=80
            nc.comment = cmt

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(details)+1}"

    ws2 = wb.create_sheet("汇总报告")
    total = len(details)
    anom  = sum(1 for d in details if d["anomaly"])
    anom_amt = sum(d["stmt_amt"] or 0 for d in details if d["anomaly"])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 36
    for ri,(k,v) in enumerate([
        ("对账任务",   task_name),
        ("生成时间",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("关联标识",   "采购单号 + SKU"),
        ("单价规则",   "含税单价有值优先；否则用单价列"),
        ("行金额规则", "系统计算：收货单数量 × 采购单单价"),
        ("",""),
        ("对账总行数", total),
        ("通过行数",   total-anom),
        ("异常行数",   anom),
        ("异常总金额", f"¥{anom_amt:,.2f}"),
        ("",""),
        ("对账结论",   "✅ 全部通过" if anom==0 else f"⚠️ 存在{anom}项异常，请复查"),
    ],1):
        ka = ws2.cell(row=ri, column=1, value=k)
        ka.font = Font(bold=True, color="1F4E79")
        va = ws2.cell(row=ri, column=2, value=v)
        if k == "对账结论":
            va.font = Font(bold=True, color="CC0000" if anom>0 else "006400")
            va.fill = RED_FILL if anom>0 else GREEN_FILL

    for sname,df in [("采购订单_原始",po_df),("收货单_原始",recv_df),("对账单_原始",stmt_df)]:
        wsr = wb.create_sheet(sname)
        for ci,col in enumerate(df.columns,1):
            c = wsr.cell(row=1, column=ci, value=str(col))
            c.fill = HEAD_FILL; c.font = HEAD_FONT
            wsr.column_dimensions[c.column_letter].width = 14
        for ri,row in enumerate(df.itertuples(index=False),2):
            for ci,val in enumerate(row,1):
                wsr.cell(row=ri, column=ci, value=val)

    wb.save(path)

